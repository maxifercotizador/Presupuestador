#!/usr/bin/env python3
"""Compara `Precio Surtidos.xlsx` (este repo) con `surtidos-data.js` (Flyers-Catalogo)
y emite un reporte en JSON con drift detectado.

Salida: si hay diferencias, exit code 0 con JSON en stdout listo para ir a un Issue;
si no hay diferencias, exit 0 con `{"diff": false}`.

Uso:
    python3 scripts/check_catalog_consistency.py \
        --excel "Precio Surtidos.xlsx" \
        --js /tmp/flyers/surtidos-data.js \
        > /tmp/drift.json
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


def parse_excel(path: str) -> dict[str, dict]:
    """Devuelve {codigo: {nombre, total_piezas, productos: [{numero, cantidad}]}}."""
    wb = load_workbook(path, data_only=True, read_only=True)
    sh = wb["BD"] if "BD" in wb.sheetnames else wb[wb.sheetnames[0]]
    out: dict[str, dict] = {}
    for i, row in enumerate(sh.iter_rows(values_only=True)):
        if i == 0:
            continue
        codigo, nombre, cant_std, num, _producto, _precio_std, *_ = list(row) + [None] * 8
        if not codigo:
            continue
        codigo = str(codigo).strip()
        if codigo not in out:
            out[codigo] = {"nombre": str(nombre).strip() if nombre else "",
                           "productos": [], "total_piezas": 0}
        try:
            qty = int(cant_std) if cant_std is not None else 0
        except (ValueError, TypeError):
            qty = 0
        out[codigo]["productos"].append({"numero": num, "cantidad": qty})
        out[codigo]["total_piezas"] += qty
    return out


def parse_js(path: str) -> dict[str, dict]:
    """Parser tolerante de surtidos-data.js. Devuelve {codigo: {nombre, piezas}}."""
    text = Path(path).read_text()
    out: dict[str, dict] = {}
    # Cada item es un bloque {...} con campos como `codigo: "X"`, `piezas: N`, `nombre: "Y"`.
    blocks = re.findall(r"\{[^{}]*\}", text)
    for b in blocks:
        m_cod = re.search(r"codigo:\s*['\"]([^'\"]+)['\"]", b)
        m_nom = re.search(r"nombre:\s*['\"]([^'\"]+)['\"]", b)
        m_pz = re.search(r"piezas:\s*(\d+)", b)
        if not m_cod:
            continue
        out[m_cod.group(1)] = {
            "nombre": m_nom.group(1) if m_nom else "",
            "piezas": int(m_pz.group(1)) if m_pz else 0,
        }
    return out


def diff(excel: dict, js: dict) -> dict:
    excel_codes = set(excel.keys())
    js_codes = set(js.keys())
    new_in_excel = sorted(excel_codes - js_codes)
    removed_from_excel = sorted(js_codes - excel_codes)
    pieza_changes = []
    for code in sorted(excel_codes & js_codes):
        ex_pz = excel[code]["total_piezas"]
        js_pz = js[code]["piezas"]
        # Solo reportar si JS tenía un número >0 y diverge — si JS está en 0 puede ser "todavía no completé"
        if ex_pz != js_pz and js_pz > 0:
            pieza_changes.append({"codigo": code,
                                  "nombre_js": js[code]["nombre"],
                                  "piezas_excel": ex_pz,
                                  "piezas_js": js_pz})
    has_diff = bool(new_in_excel or removed_from_excel or pieza_changes)
    return {
        "diff": has_diff,
        "new_in_excel": [{"codigo": c, "nombre_excel": excel[c]["nombre"],
                          "total_piezas": excel[c]["total_piezas"]} for c in new_in_excel],
        "removed_from_excel": [{"codigo": c, "nombre_js": js[c]["nombre"]}
                               for c in removed_from_excel],
        "pieza_changes": pieza_changes,
        "totals": {"excel": len(excel), "js": len(js)},
    }


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel", required=True)
    ap.add_argument("--js", required=True)
    args = ap.parse_args()
    excel = parse_excel(args.excel)
    js = parse_js(args.js)
    result = diff(excel, js)
    print(json.dumps(result, indent=2, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    sys.exit(main())
